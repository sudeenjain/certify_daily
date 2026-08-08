#!/usr/bin/env python3
"""
Daily certification/course update monitor -> structured Excel digest.

For each source in sources.json:
  - RSS sources: pulls the latest entry title + link + summary text.
  - HTML sources: hashes the page content and compares to the last
    saved hash to detect whether it changed since yesterday.

For every source that has something new today (a changed HTML page,
or a fresh RSS entry), the page/entry text is sent to Claude, which
extracts these fields as best it can find them on the page:
  Name, Dept, Direct Register Link, Requirement, Announced/Open Date,
  Start Date, End Date, Badge (Free/Paid/Not offered/Not found),
  Certificate (Free/Paid/Not offered/Not found), Provider, Status,
  Description

Not every source page states all of these clearly - fields the page
doesn't mention come back as "Not specified". This is a best-effort
extraction, not a guarantee of accuracy - always verify on the
provider's own site before relying on dates or pricing.

Output: an .xlsx file attached to the daily email, plus a short plain
text summary in the email body. Sources with no change today are
skipped from the sheet entirely to keep it focused.

State (previous hashes) is persisted in state.json, which this script
rewrites each run. In GitHub Actions, the workflow commits state.json
back to the repo so history carries over between runs.
"""

import json
import os
import re
import smtplib
import logging
import hashlib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime

import requests
import feedparser
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
log = logging.getLogger(__name__)

SOURCES_FILE = "sources.json"
STATE_FILE = "state.json"
DIGEST_XLSX = "digest.xlsx"
REQUEST_TIMEOUT = 20
HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; CertMonitorBot/1.0; +https://github.com/)"
}
ANTHROPIC_MODEL = "claude-haiku-4-5-20251001"  # cheap + fast, good enough for extraction
COLUMNS = [
    "Name", "Dept", "Direct Register Link", "Requirement",
    "Announced/Open Date", "Start Date", "End Date",
    "Badge", "Certificate", "Provider", "Status", "Description",
]


def load_json(path, default):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return default


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def normalize_html_text(html):
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg", "footer", "nav"]):
        tag.decompose()
    text = soup.get_text(separator=" ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def check_rss(source):
    try:
        feed = feedparser.parse(source["url"])
        if not feed.entries:
            return {"status": "error", "detail": "No entries found in feed"}
        latest = feed.entries[0]
        return {
            "status": "ok",
            "latest_title": latest.get("title", "(no title)"),
            "latest_link": latest.get("link", source["url"]),
            "summary": BeautifulSoup(latest.get("summary", ""), "html.parser").get_text()[:2000],
            "published": latest.get("published", ""),
        }
    except Exception as e:
        return {"status": "error", "detail": str(e)}


def check_html(source, prev_hash):
    try:
        resp = requests.get(source["url"], headers=HEADERS, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        text = normalize_html_text(resp.text)
        digest = hashlib.sha256(text.encode("utf-8")).hexdigest()
        changed = prev_hash is not None and digest != prev_hash
        first_check = prev_hash is None
        return {
            "status": "ok",
            "hash": digest,
            "changed": changed,
            "first_check": first_check,
            "text": text[:6000],  # cap what we send to the LLM
        }
    except Exception as e:
        return {"status": "error", "detail": str(e)}


def clean_env(val):
    if val is None:
        return None
    return val.replace("\xa0", "").strip()


def extract_fields_with_claude(name, url, text_snippet, api_key):
    """Ask Claude to pull structured fields out of the page/entry text.
    Falls back to 'Not specified' fields on any failure so one bad
    extraction never breaks the whole run."""
    fallback = {
        "Dept": "Not specified", "Requirement": "Not specified",
        "Announced/Open Date": "Not specified", "Start Date": "Not specified",
        "End Date": "Not specified", "Badge": "Not specified",
        "Certificate": "Not specified", "Status": "See link",
        "Description": text_snippet[:200].strip() or "No summary available",
    }
    if not api_key:
        return fallback

    prompt = f"""You are extracting structured info from a certification/course provider's web page or feed entry, for a personal tracking spreadsheet.

Source name: {name}
URL: {url}
Page/entry text (truncated):
---
{text_snippet[:5000]}
---

Return ONLY a JSON object (no markdown, no preamble) with exactly these keys:
- "Dept": the department/category/track this belongs to if stated (e.g. "Cloud", "Data", "Security"), else "Not specified"
- "Requirement": prerequisites to take it if stated, else "Not specified"
- "Announced/Open Date": date registration opened/was announced if stated, else "Not specified"
- "Start Date": course/cohort start date if stated, else "Not specified"
- "End Date": course/cohort end date or registration deadline if stated, else "Not specified"
- "Badge": one of "Free", "Paid", "Not offered", "Not specified" - whether a digital badge is offered and its cost
- "Certificate": one of "Free", "Paid", "Not offered", "Not specified" - whether a certificate is offered and its cost
- "Status": one of "New", "Updated", "Open", "Closed", "Not specified" - your best read of the current status from the text
- "Description": a plain one-sentence summary of what this page/post is actually about, in your own words, under 25 words

Only state a field if the text actually supports it. Do not guess or invent dates or prices."""

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": ANTHROPIC_MODEL,
                "max_tokens": 500,
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        text = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")
        text = text.strip()
        text = re.sub(r"^```(json)?|```$", "", text, flags=re.MULTILINE).strip()
        parsed = json.loads(text)
        result = dict(fallback)
        result.update({k: v for k, v in parsed.items() if k in fallback})
        return result
    except Exception as e:
        log.warning(f"Claude extraction failed for {name}: {e}")
        return fallback


def build_rows(results, api_key):
    rows = []
    for r in results:
        name, url, res = r["name"], r["url"], r["result"]

        if r["type"] == "html":
            if res.get("status") != "ok":
                continue  # failed checks don't get a spreadsheet row
            if res.get("first_check"):
                continue  # no "yesterday" to compare - nothing to report yet
            if not res.get("changed"):
                continue  # unchanged - skip, keep sheet focused
            snippet = res.get("text", "")
            log.info(f"Extracting fields for changed page: {name}")
            fields = extract_fields_with_claude(name, url, snippet, api_key)

        elif r["type"] == "rss":
            if res.get("status") != "ok":
                continue
            snippet = f"{res['latest_title']}\n{res.get('summary','')}"
            log.info(f"Extracting fields for RSS item: {name}")
            fields = extract_fields_with_claude(name, url, snippet, api_key)
            url = res["latest_link"]
        else:
            continue

        rows.append({
            "Name": name,
            "Dept": fields["Dept"],
            "Direct Register Link": url,
            "Requirement": fields["Requirement"],
            "Announced/Open Date": fields["Announced/Open Date"],
            "Start Date": fields["Start Date"],
            "End Date": fields["End Date"],
            "Badge": fields["Badge"],
            "Certificate": fields["Certificate"],
            "Provider": name,
            "Status": fields["Status"],
            "Description": fields["Description"],
        })
    return rows


def build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Certification Updates"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [22, 14, 32, 22, 16, 14, 14, 10, 10, 22, 12, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(rows)+1,1)}"

    wb.save(DIGEST_XLSX)
    return DIGEST_XLSX


def update_google_sheet(rows, run_date):
    """Append today's rows to a Google Sheet, creating the header if the
    sheet is empty. Silently skips (with a log line) if credentials or
    the sheet ID aren't configured - this feature is optional."""
    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    sheet_id = os.environ.get("GOOGLE_SHEET_ID")
    if not sa_json or not sheet_id:
        log.info("GOOGLE_SERVICE_ACCOUNT_JSON / GOOGLE_SHEET_ID not set - skipping Google Sheet update")
        return

    if not rows:
        log.info("No new rows today - nothing to append to Google Sheet")
        return

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds_info = json.loads(sa_json)
        creds = Credentials.from_service_account_info(creds_info, scopes=scopes)
        gc = gspread.authorize(creds)

        sh = gc.open_by_key(sheet_id)
        sheet_name = os.environ.get("GOOGLE_SHEET_TAB", "Certification Updates")
        try:
            ws = sh.worksheet(sheet_name)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=sheet_name, rows=1000, cols=len(COLUMNS) + 1)

        existing = ws.get_all_values()
        if not existing:
            ws.append_row(["Date Found"] + COLUMNS, value_input_option="USER_ENTERED")

        new_rows = [[run_date] + [row.get(col, "") for col in COLUMNS] for row in rows]
        ws.append_rows(new_rows, value_input_option="USER_ENTERED")
        log.info(f"Appended {len(new_rows)} row(s) to Google Sheet '{sheet_name}'")
    except Exception as e:
        log.warning(f"Google Sheet update failed: {e}")


def send_email(subject, body, attachment_path, errors):
    smtp_user = clean_env(os.environ.get("SMTP_USER"))
    smtp_pass = clean_env(os.environ.get("SMTP_PASS"))
    to_addr = clean_env(os.environ.get("TO_EMAIL")) or smtp_user

    if not smtp_user or not smtp_pass:
        log.error("SMTP_USER / SMTP_PASS not set - printing digest instead")
        print(body)
        return

    msg = MIMEMultipart()
    msg["From"] = smtp_user
    msg["To"] = to_addr
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    if attachment_path and os.path.exists(attachment_path):
        with open(attachment_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(attachment_path)}")
        msg.attach(part)

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_user, [to_addr], msg.as_string())
    log.info(f"Email sent to {to_addr}")


def main():
    sources = load_json(SOURCES_FILE, [])
    state = load_json(STATE_FILE, {})
    run_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    api_key = clean_env(os.environ.get("ANTHROPIC_API_KEY"))

    if not api_key:
        log.warning("ANTHROPIC_API_KEY not set - spreadsheet fields will be 'Not specified' instead of extracted")

    results = []
    for source in sources:
        name = source["name"]
        stype = source["type"]
        log.info(f"Checking: {name} ({stype})...")

        if stype == "rss":
            res = check_rss(source)
        else:
            prev_hash = state.get(name, {}).get("hash")
            res = check_html(source, prev_hash)
            if res.get("status") == "ok":
                state[name] = {"hash": res["hash"], "last_checked": run_date}

        results.append({"name": name, "url": source["url"], "type": stype, "result": res})

    save_json(STATE_FILE, state)

    rows = build_rows(results, api_key)
    xlsx_path = build_workbook(rows)
    update_google_sheet(rows, run_date)

    errors = [r for r in results if r["result"].get("status") == "error"]
    body_lines = [
        f"Certification / course update digest - {run_date}",
        "",
        f"{len(rows)} update(s) found today - see attached spreadsheet for full details.",
        "",
    ]
    if errors:
        body_lines.append(f"{len(errors)} source(s) failed to check today (spreadsheet unaffected, listed below):")
        for r in errors:
            body_lines.append(f"  - {r['name']}: {r['result'].get('detail','unknown error')}")
        body_lines.append("")
    body_lines.append(
        "Note: fields are extracted automatically from each page's text and may say "
        "'Not specified' where the page doesn't state that detail. Always confirm exact "
        "dates and pricing on the provider's own site before registering."
    )
    body = "\n".join(body_lines)

    subject = f"Daily Certification Update Digest - {datetime.now().strftime('%d %b %Y')} ({len(rows)} updates)"
    send_email(subject, body, xlsx_path, errors)
    log.info("Monitor run complete")


if __name__ == "__main__":
    main()
